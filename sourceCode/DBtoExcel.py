
import pymysql
import numpy as np
import argparse
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook

# Parse command line arguments
parser = argparse.ArgumentParser(description="Execute SQL queries on a database.")
parser.add_argument("database", type=str, help="Name of the database to execute queries on.")
parser.add_argument("ip_address", type=str, help="ip_address")
parser.add_argument("runningtime", type=str, help="Time to run script")
parser.add_argument("objects", type=str, help="Objects in the frame")
parser.add_argument("maindbhost", type=str, help="Host of the main database server")

args = parser.parse_args()
db_name = args.database
ip_address = args.ip_address
time_to_run=args.runningtime
objects=args.objects
main_db_host = args.maindbhost

# Connect to the database
try:
    conn = pymysql.connect(
        host=main_db_host,
        user='root',
        password='pass',
        db=db_name,
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )
except pymysql.Error as e:
    print(f"Error connecting to database: {e}")
    exit()
    # Handle the error as appropriate

# Execute the query to fetch the lowest and highest timeStamp values
try:
    with conn.cursor() as cursor:
        cursor.execute("""
            SELECT MIN(TimeStamp) AS min_timestamp, MAX(TimeStamp) AS max_timestamp
            FROM camerainfo
        """)
        result = cursor.fetchone()
        min_timestamp = result['min_timestamp']
        max_timestamp = result['max_timestamp']
except pymysql.Error as e:
    print(f"Error executing SQL query: {e}")
    # Handle the error as appropriate

# Execute the query to fetch the lowest and highest timeStamp values
try:
    with conn.cursor() as cursor:
        cursor.execute("""
            SELECT RelativeTime FROM camerainfo LIMIT 1
        """)
        result = cursor.fetchone()
        firstRelativeTime = result['RelativeTime']
except pymysql.Error as e:
    print(f"Error executing SQL query: {e}")
    # Handle the error as appropriate

#TargetFPS
try:
    with conn.cursor() as cursor:
        sql = "SELECT TargetFPS FROM serverinfo WHERE ServerIP = %s"
        cursor.execute(sql,(ip_address))
        rows = cursor.fetchall()
        targetFPS=rows[0]["TargetFPS"]
except pymysql.Error as e:
    print(f"MySQL query error: {e}")


# Parse the time string into a datetime object
time_obj = datetime.strptime(firstRelativeTime, "%H:%M:%S")

# Convert the datetime object to seconds
seconds = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second


# Reduce 10 seconds from the min timestamp
min_timestamp = datetime.strptime(str(min_timestamp), '%Y-%m-%d %H:%M:%S')
min_timestamp -= timedelta(seconds=seconds)
max_timestamp = min_timestamp + timedelta(seconds=int(time_to_run))
min_timestamp = min_timestamp.strftime('%Y-%m-%d %H:%M:%S')

# Add 10 seconds to the max timestamp
# max_timestamp = datetime.strptime(str(max_timestamp), '%Y-%m-%d %H:%M:%S')
# max_timestamp = min_timestamp + timedelta(seconds=time_to_run)
max_timestamp = max_timestamp.strftime('%Y-%m-%d %H:%M:%S')

print("Start time :",min_timestamp)
print("Stop time  :",max_timestamp)
# exit()

# Execute the query to fetch all unique camera names
try:
    with conn.cursor() as cursor:
        cursor.execute("""
            SELECT DISTINCT CameraName
            FROM camerainfo
        """)
        results = cursor.fetchall()
        camera_names = [result['CameraName'] for result in results]
except pymysql.Error as e:
    print(f"Error executing SQL query: {e}")
    # Handle the error as appropriate

camera_names = sorted(camera_names)

camNames = f"{camera_names[0]}-{camera_names[-1]}"
# print(camNames)  # Output: LOAD1-5



try:
    with conn.cursor() as cursor:
        sql = "SELECT Features FROM serverinfo WHERE ServerIP = %s"
        cursor.execute(sql,(ip_address))
        rows = cursor.fetchall()
        features_str=rows[0]["Features"]
except pymysql.Error as e:
    print(f"MySQL query error: {e}")



try:
    cursor = conn.cursor()
    # get average analytics, OC,ANPR,FR utilization from cpuusage table

    # Get the column names from the table
    cursor.execute("SHOW COLUMNS FROM cpumemoryusage")
    columns = cursor.fetchall()
    column_names = [column['Field'] for column in columns]
    #print(column_names)

    if 'analytics' in column_names:
        sql = "SELECT AVG(analytics) FROM cpuusage where TimeStamp >= %s and TimeStamp < %s"
        cursor = conn.cursor()
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avganalyticsutil = cursor.fetchall()
        # print(avganalyticsutil)
        avganalyticsutil = avganalyticsutil[0]['AVG(analytics)'] if avganalyticsutil and avganalyticsutil[0]['AVG(analytics)'] else 0

        sql = "SELECT AVG(analytics) FROM cpumemoryusage where TimeStamp >= %s and TimeStamp < %s"
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avganalyticsmemory = cursor.fetchall()
        avganalyticsmemory = avganalyticsmemory[0]['AVG(analytics)'] if avganalyticsmemory and avganalyticsmemory[0]['AVG(analytics)'] else 0
    
    else:
        avganalyticsmemory=0
        avganalyticsutil=0


    if 'OC' in column_names:
        sql = "SELECT AVG(OC)  FROM cpuusage where TimeStamp >= %s and TimeStamp < %s"
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avgocutil = cursor.fetchall()
        # print(avgocutil)
        avgocutil = avgocutil[0]['AVG(OC)'] if avgocutil and avgocutil[0]['AVG(OC)'] else 0
        
        sql = "SELECT AVG(OC)  FROM cpumemoryusage where TimeStamp >= %s and TimeStamp < %s"
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avgocmemory = cursor.fetchall()
        avgocmemory = avgocmemory[0]['AVG(OC)'] if avgocmemory and avgocmemory[0]['AVG(OC)'] else 0
    else:
        avgocmemory=0
        avgocutil=0  

    if 'TS' in column_names:
        sql = "SELECT AVG(TS) FROM cpuusage where TimeStamp >= %s and TimeStamp < %s"
        cursor = conn.cursor()
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avgtsutil = cursor.fetchall()
        # print(avganalyticsutil)
        avgtsutil = avgtsutil[0]['AVG(TS)'] if avgtsutil and avgtsutil[0]['AVG(TS)'] else 0

        sql = "SELECT AVG(TS) FROM cpumemoryusage where TimeStamp >= %s and TimeStamp < %s"
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avgtsmemory = cursor.fetchall()
        avgtsmemory = avgtsmemory[0]['AVG(TS)'] if avgtsmemory and avgtsmemory[0]['AVG(TS)'] else 0
    
    else:
        avgtsmemory=0
        avgtsutil=0       

    if 'ANPR' in column_names:
        sql = "SELECT AVG(ANPR) FROM cpuusage where TimeStamp >= %s and TimeStamp < %s"
        cursor = conn.cursor()
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avganprutil = cursor.fetchall()
        #print(avganprutil)
        avganprutil = avganprutil[0]['AVG(ANPR)'] if avganprutil and avganprutil[0]['AVG(ANPR)'] else 0
        
        sql = "SELECT AVG(ANPR)  FROM cpumemoryusage where TimeStamp >= %s and TimeStamp < %s"
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avganprmemory = cursor.fetchall()
        avganprmemory = avganprmemory[0]['AVG(ANPR)'] if avganprmemory and avganprmemory[0]['AVG(ANPR)'] else 0
        #print(avganprmemory)
    else:
        avganprmemory=0
        avganprutil=0

    # Check if 'FR' column exists in the table
    if 'FR' in column_names:
        # Column exists, execute the query
        sql = "SELECT AVG(FR) FROM cpumemoryusage WHERE TimeStamp >= %s AND TimeStamp < %s"
        cursor.execute(sql, (min_timestamp, max_timestamp))
        avgfrmemory = cursor.fetchall()
        avgfrmemory = avgfrmemory[0]['AVG(FR)'] if avgfrmemory and avgfrmemory[0]['AVG(FR)'] else 0
        
        sql = "SELECT AVG(FR)  FROM cpuusage where TimeStamp >= %s and TimeStamp < %s"
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avgfrutil = cursor.fetchall()
        # print(avgfrutil)
        avgfrutil = avgfrutil[0]['AVG(FR)'] if avgfrutil and avgfrutil[0]['AVG(FR)'] else 0

    else:
        # Column does not exist
        avgfrmemory = 0
        avgfrutil=0
       
    # Check if 'Decoder' column exists in the table
    if 'decode_service' in column_names:
        # Column exists, execute the query
        sql = "SELECT AVG(decode_service) FROM cpumemoryusage WHERE TimeStamp >= %s AND TimeStamp < %s"
        cursor.execute(sql, (min_timestamp, max_timestamp))
        avgdecodememory = cursor.fetchall()
        avgdecodememory = avgdecodememory[0]['AVG(decode_service)'] if avgdecodememory and avgdecodememory[0]['AVG(decode_service)'] else 0
        
        sql = "SELECT AVG(decode_service)  FROM cpuusage where TimeStamp >= %s and TimeStamp < %s"
        cursor.execute(sql,(min_timestamp,max_timestamp))
        avgdecodeutil = cursor.fetchall()
        #print(avgdecodeutil)
        # print(avgfrutil)
        avgdecodeutil = avgdecodeutil[0]['AVG(decode_service)'] if avgdecodeutil and avgdecodeutil[0]['AVG(decode_service)'] else 0

    else:
        # Column does not exist
        avgdecodememory = 0
        avgdecodeutil=0 
          

    # get GPU power usage, percentage and memory usage from gpustats table
    gpu_power_usage = -1
    query = "SELECT AVG(CurrentPowerUsage) FROM gpustats where TimeStamp >=%s and TimeStamp < %s"
    cursor.execute(query,(min_timestamp,max_timestamp))
    y = cursor.fetchall()
    if y and y[0]['AVG(CurrentPowerUsage)']:
        gpu_power_usage = y[0]['AVG(CurrentPowerUsage)']

    gpu_percent = -1
    query = "SELECT AVG(GPUUsagePerc) FROM gpustats where TimeStamp >=%s and TimeStamp < %s"
    cursor.execute(query,(min_timestamp,max_timestamp))
    y = cursor.fetchall()
    # print(y)
    if y:
        # print("Here")
        gpu_percent = y[0]['AVG(GPUUsagePerc)']

    gpu_memory = -1
    query = "SELECT AVG(GPUMemory) FROM gpustats where TimeStamp >=%s and TimeStamp < %s"
    cursor.execute(query,(min_timestamp,max_timestamp))
    y = cursor.fetchall()
    if y and y[0]['AVG(GPUMemory)']:
        gpu_memory = y[0]['AVG(GPUMemory)']

    # get average input and output fps from camerainfo table
    avg_input_fps = -1
    avg_output_fps = -1
    query = "SELECT AVG(InputFPS) FROM camerainfo where TimeStamp >=%s and TimeStamp < %s"
    cursor.execute(query,(min_timestamp,max_timestamp))
    y = cursor.fetchall()
    if y and y[0]['AVG(InputFPS)']:
        avg_input_fps = y[0]['AVG(InputFPS)']

    query = "SELECT AVG(OutputFPS) FROM camerainfo where TimeStamp >=%s and TimeStamp < %s"
    cursor.execute(query,(min_timestamp,max_timestamp))
    y = cursor.fetchall()
    if y and y[0]['AVG(OutputFPS)']:
        avg_output_fps = y[0]['AVG(OutputFPS)']

except Exception as e:
    print("Error: ", e)


output_fps_90thpercentile = -1
query = "SELECT OutputFPS FROM camerainfo WHERE TimeStamp >= %s AND TimeStamp < %s"
cursor = conn.cursor()
try:
    cursor.execute(query, (min_timestamp, max_timestamp))
    rows = cursor.fetchall()
    # print(rows)
    data = [row['OutputFPS'] for row in rows] # Extract the OutputFPS values from the fetched rows
    output_fps_90thpercentile = np.percentile(data, 90)
except Exception as e:
    print("Error fetching 10th percentile output fps:", e)

# ram
ram = -1
query = "SELECT RAM FROM serverinfo"
cursor = conn.cursor()
try:
    cursor.execute(query)
    y = cursor.fetchall()
    # print(y)
    ram = y[0]['RAM']
except Exception as e:
    print("Error fetching RAM:", e)

# cores
cores = -1
query = "SELECT CPUcores FROM serverinfo"
cursor = conn.cursor()
try:
    cursor.execute(query)
    y = cursor.fetchall()
    cores = y[0]['CPUcores']
except Exception as e:
    print("Error fetching CPU cores:", e)

# gpuname and modelname
gpuname = -1
query = "SELECT GPUName FROM serverinfo where ServerIP=%s"
cursor = conn.cursor()
try:
    cursor.execute(query,ip_address)
    y = cursor.fetchall()
    gpuname = y[0]['GPUName']
except Exception as e:
    print("Error fetching GPU name:", e)

modelname = -1
query = "SELECT  ModelName FROM serverinfo where ServerIP=%s"
cursor = conn.cursor()
try:
    cursor.execute(query,ip_address)
    y = cursor.fetchall()
    modelname = y[0]['ModelName']
except Exception as e:
    print("Error fetching model name:", e)

#****decoder_service_info*****


#acrossallcameras
decode_avg_fps=0
decode_90thpercentile_fps = 0
query = "SELECT AVG(AvgDecodeFPS) FROM decoderserviceinfo where TimeStamp >=%s and TimeStamp < %s"
cursor.execute(query,(min_timestamp,max_timestamp))

y = cursor.fetchall()
#print(y)
if y and y[0]['AVG(AvgDecodeFPS)']:
  decode_avg_fps = y[0]['AVG(AvgDecodeFPS)']
  #print(decode_fps)
else:
   decode_avg_fps=0  
# Execute count query
if decode_avg_fps!=0:
    count_query = "SELECT COUNT(*) FROM decoderserviceinfo WHERE TimeStamp >= %s AND TimeStamp < %s"
    cursor.execute(count_query, (min_timestamp, max_timestamp))
    count_result = cursor.fetchone()
    # Extract row count value
    #print(count_result)
    row_count = count_result['COUNT(*)'] if count_result else 0
    #print(row_count)
    

    #90thpercentileacrossall
    query = "SELECT AvgDecodeFPS FROM decoderserviceinfo WHERE TimeStamp >= %s AND TimeStamp < %s"
    cursor = conn.cursor()
    try:
           cursor.execute(query, (min_timestamp, max_timestamp))
           rows = cursor.fetchall()
           #print(rows)
           # print(rows)
           data = [float(row['AvgDecodeFPS']) for row in rows] # Extract the OutputFPS values from the fetched rows
           #print(data)
           #print("**")
           decode_90thpercentile_fps = np.percentile(data, 90)
           #print(decode_90thpercentile_fps)
           #print("**")
    except Exception as e:
           print("Error fetching 10th percentile decode fps:", e)


#across_cpu

decode_cpu_avg_fps =0
decode_90thpercentile_fps_cpu=0
query = "SELECT AVG(AvgDecodeFPS) FROM decoderserviceinfo where TimeStamp >=%s and TimeStamp < %s and HardwareType='cpu'"
cursor.execute(query,(min_timestamp,max_timestamp))
y = cursor.fetchall()
#print(y)
if y and y[0]['AVG(AvgDecodeFPS)']:
  decode_cpu_avg_fps = y[0]['AVG(AvgDecodeFPS)']
else:
   decode_cpu_avg_fps=0
if(decode_cpu_avg_fps!=0):     
   # Execute count query
   count_query = "SELECT COUNT(*) FROM decoderserviceinfo WHERE TimeStamp >= %s AND TimeStamp < %s and HardwareType='cpu'"
   cursor.execute(count_query, (min_timestamp, max_timestamp))
   count_result = cursor.fetchone()
   #print(count_result)
   # Extract row count value
   row_count = count_result['COUNT(*)'] if count_result else 0
   #print(row_count)
   #print(decode_cpu_fps)

   #90th percentile across cpu
   query = "SELECT AvgDecodeFPS FROM decoderserviceinfo WHERE TimeStamp >= %s AND TimeStamp < %s and HardwareType='cpu'"
   cursor = conn.cursor()
   try:
      cursor.execute(query, (min_timestamp, max_timestamp))
      rows = cursor.fetchall()
      # print(rows)
      data = [float(row['AvgDecodeFPS']) for row in rows] # Extract the OutputFPS values from the fetched rows
      decode_90thpercentile_fps_cpu = np.percentile(data, 90)
   except Exception as e:
      print("Error fetching 10th percentile decode fps:", e)


#across gpu

decode_gpu_avg_fps =0
decode_90thpercentile_fps_gpu=0
query = "SELECT AVG(AvgDecodeFPS) FROM decoderserviceinfo where TimeStamp >=%s and TimeStamp < %s and HardwareType='gpu'"
cursor.execute(query,(min_timestamp,max_timestamp))
y = cursor.fetchall()
if y and y[0]['AVG(AvgDecodeFPS)']:
  decode_gpu_avg_fps = y[0]['AVG(AvgDecodeFPS)']
  #print(decode_gpu_fps)
  #print("ash") 
else:
    decode_gpu_avg_fps=0 

if (decode_gpu_avg_fps!=0):
      # Execute count query
      count_query = "SELECT COUNT(*) FROM decoderserviceinfo WHERE TimeStamp >= %s AND TimeStamp < %s and HardwareType='gpu'"
      cursor.execute(count_query, (min_timestamp, max_timestamp))
      count_result = cursor.fetchone()
      # Extract row count value
      #print(count_result)
      row_count = count_result['COUNT(*)'] if count_result else 0
      #print(row_count)

      #90thpercentileacrossgpu
      query = "SELECT AvgDecodeFPS FROM decoderserviceinfo WHERE TimeStamp >= %s AND TimeStamp < %s and HardwareType='gpu'"
      cursor = conn.cursor()
      try:
             cursor.execute(query, (min_timestamp, max_timestamp))
             rows = cursor.fetchall()
             # print(rows)
             data = [float(row['AvgDecodeFPS']) for row in rows] # Extract the OutputFPS values from the fetched rows
             decode_90thpercentile_fps_gpu = np.percentile(data, 90)
      except Exception as e:
            print("Error fetching 10th percentile decode fps:", e)

# print("decode_avg_fps",decode_avg_fps)
# print("decode_90thpercentile_fps",decode_90thpercentile_fps)
# print("decode_cpu_avg_fps",decode_cpu_avg_fps)
# print("decode_90thpercentile_fps_cpu",decode_90thpercentile_fps_cpu)
# print("decode_gpu_avg_fps",decode_gpu_avg_fps)  
# print("decode_90thpercentile_fps_gpu",decode_90thpercentile_fps_gpu)







data={}
# data['SlNo']=1
data["testcaseID"]=db_name
data["Scenarios"]="0"
data["Features"]=features_str

# Path to config.json file
file_path = '/home/allgovision/AGV/data/AllGoVision/ProgramData/public/configuration/config.json'

# Read JSON data from file and parse it into a dictionary
with open(file_path, 'r') as file:
    parsed_data = json.load(file)

# Get service version
service_version = parsed_data["serviceVersion"]

# Assign service version to "Web Application" key in data dictionary
data["Web Application"] = str(service_version)
data["Server"]=ip_address
data["Camera Name"]=camNames
data["No of cams"]=len(camera_names)
data['analytics_core_util'] = float(avganalyticsutil)/int(cores)
data['analytics_mem_util'] = float(avganalyticsmemory)/1024.0
#print(data['analytics_mem_util'])

data['TS_core_util']=float(avgtsutil)/int(cores)
data['TS_mem_util']=float(avgtsmemory)/1024.0
data['OC_core_util']=float(avgocutil)/int(cores)
data['OC_mem_util']=float(avgocmemory)/1024.0
data['ANPR_core_util']=float(avganprutil)/int(cores)
data['ANPR_mem_util']=float(avganprmemory)/1024.0
#print(avgfrutil)
data['FR_core_util']=float(avgfrutil)/float(cores)
data['FR_mem_util']=float(avgfrmemory)/1024.0
#data['FR_core_util']=-1
#data['FR_mem_util']=-1
data['decoder_core_util']=float(avgdecodeutil)/float(cores)
data['decoder_mem_util']=float(avgdecodememory)/1024.0
    


data["Total CPU%"]=data["analytics_core_util"]+data["OC_core_util"]+data["ANPR_core_util"]+data["FR_core_util"]+data['decoder_core_util']+data['TS_core_util']
data["Total memory(GB)"]=data["analytics_mem_util"]+data["OC_mem_util"]+data["ANPR_mem_util"]+data["FR_mem_util"]+data['decoder_mem_util']+data['TS_mem_util']
data["GPU-Power"]=float(gpu_power_usage)
data["GPU%"]=float(gpu_percent)
data["GPU-Memory"]=float(gpu_memory)/1024.0
data["Analytics Resolution"]=1080
data["Input fps"]=25
data["Target fps"]=targetFPS
data['Average input FPS']=float(avg_input_fps)
data['Average output FPS']=float(avg_output_fps)    
data['90th percentile']=float(output_fps_90thpercentile)
data['decoder_avg_fps']=decode_avg_fps
data['decoder_90thpercentile_fps']=decode_90thpercentile_fps
data['cpu_avg_fps']=decode_cpu_avg_fps
data['decoder_cpu_90thpercentile_fps']=decode_90thpercentile_fps_cpu
data['decoder_gpu_avg_fps']=decode_gpu_avg_fps
data['decoder_gpu_90thpercentile']=decode_90thpercentile_fps_gpu
data["Ram(GB)"]=ram #free -h
data["Cores"]=cores #lscpu -B
data["GOP Length"]=150
data["Bit rate"]=1359
data["image"]="No"
data["metadata"]="No"
data["No of core per channel"]=(data["Total CPU%"]*float(cores))/(100.0*float(data["No of cams"]))
data["MHz per channel"]=data["No of core per channel"]*2000.0
data["MHz per output image"]=data["MHz per channel"]/float(data['Average output FPS'])
data["Objects in frame"]=objects
data["Nvidia"]=gpuname#nvidia-smi q
data["Model"]=modelname #lscpu -B
data["DL/Non DL"]="DL"
data["Server Settings"]="-"
data["Status"]="-"

sheetpath = "Book.xlsx"
wb = load_workbook(sheetpath)
sheet = wb['Sheet1']

# Find the first empty row starting from the 6th row
row_number = 6
while sheet.cell(row=row_number, column=1).value is not None:
    row_number += 1

# Find the last filled row to determine the next serial number
last_row_number = row_number - 1
last_serial_number = sheet.cell(row=last_row_number, column=1).value
if last_serial_number == "Sl.No":
    next_serial_number = 1
else:
    next_serial_number = int(last_serial_number) + 1 

# Write the serial number in column 1 of the new row
sheet.cell(row=row_number, column=1).value = next_serial_number

for k in range(len(data)):
    sheet.cell(row=row_number, column=1+k+1).value = data[list(data.keys())[k]]

wb.save(sheetpath)
wb.close()
